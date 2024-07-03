import requests
from bs4 import BeautifulSoup
import zipfile
import os
import subprocess
import pandas as pd
import re
from datetime import datetime
import logging
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')

# Directory setup
base_dir = os.path.abspath('bnf_snomed_mapping_data')
dirs = {
    'download': base_dir,
    'latest': os.path.join(base_dir, 'latest'),
    'csv': os.path.join(base_dir, 'csv_files'),
    'xlsx': os.path.join(base_dir, 'xlsx_files'),
    'zip': os.path.join(base_dir, 'zip_files'),
    'output': os.path.join(base_dir, 'output')
}
for d in dirs.values():
    os.makedirs(d, exist_ok=True)

# Step 1: Navigate to the webpage and find all .zip hyperlinks
url = "https://www.nhsbsa.nhs.uk/prescription-data/understanding-our-data/bnf-snomed-mapping"
response = requests.get(url)
soup = BeautifulSoup(response.content, 'html.parser')

# Step 2: Download and decompress the .zip files containing .xlsx files
zip_links = [a['href'] for a in soup.find_all('a', href=True) if a['href'].endswith('.zip')]
logging.info(f"Found {len(zip_links)} zip files to download.")

xlsx_files = []

for link in zip_links:
    zip_url = link if link.startswith('http') else f"https://www.nhsbsa.nhs.uk{link}"
    zip_filename = zip_url.split('/')[-1]
    local_zip_path = os.path.join(dirs['zip'], zip_filename)
    
    # Check if the zip file is already downloaded
    if not os.path.exists(local_zip_path):
        logging.info(f"Downloading {zip_url}")
        zip_response = requests.get(zip_url)
        with open(local_zip_path, 'wb') as f:
            f.write(zip_response.content)
    else:
        logging.info(f"{zip_filename} already exists, skipping download.")
    
    with zipfile.ZipFile(local_zip_path, 'r') as zip_file:
        for file_info in zip_file.infolist():
            if file_info.filename.endswith('.xlsx'):
                extracted_path = os.path.join(dirs['xlsx'], os.path.basename(file_info.filename))
                if not os.path.exists(extracted_path):
                    logging.info(f"Extracting {file_info.filename}")
                    with zip_file.open(file_info.filename) as source, open(extracted_path, 'wb') as target:
                        target.write(source.read())
                xlsx_files.append(extracted_path)

# Step 3: Convert .xlsx to .csv using VBA if necessary
def convert_xlsx_to_csv(xlsx_file, csv_file):
    vbs_script = os.path.abspath("convert_to_csv.vbs")
    result = subprocess.run(['cscript', '//nologo', vbs_script, xlsx_file, csv_file], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    
    if result.returncode != 0:
        logging.error(f"Error during conversion: {result.stderr.decode('utf-8')}")
        logging.info(f"VBA script stdout: {result.stdout.decode('utf-8')}")
        return False
    elif result.stdout or result.stderr:
        logging.info(f"VBA script stdout: {result.stdout.decode('utf-8')}")
        logging.info(f"VBA script stderr: {result.stderr.decode('utf-8')}")
        
    return os.path.exists(csv_file)

# Step 4: Read .xlsx files, convert to .csv if necessary, and combine data into a single dataframe with distinct entries
unique_entries = set()
combined_rows = []

# Process all files
for xlsx_file in xlsx_files:
    csv_file = os.path.join(dirs['csv'], os.path.basename(xlsx_file).replace('.xlsx', '.csv'))

    # Convert .xlsx to .csv if it doesn't already exist
    if not os.path.exists(csv_file):
        logging.info(f"Converting {os.path.basename(xlsx_file)} to {os.path.basename(csv_file)}")
        if not convert_xlsx_to_csv(xlsx_file, csv_file):
            logging.error(f"Failed to convert {os.path.basename(xlsx_file)} to {os.path.basename(csv_file)}")
            continue
    else:
        logging.info(f"{os.path.basename(csv_file)} already exists, skipping conversion.")

    try:
        logging.info(f"Reading {os.path.basename(csv_file)} using pandas")
        file_unique_count = 0
        for chunk in pd.read_csv(csv_file, chunksize=10000, usecols=range(11), dtype=str):  # Read only columns A:K
            chunk.fillna('', inplace=True)  # Ensure that NaN values are replaced with empty strings
            for index, row in chunk.iterrows():
                row_tuple = tuple(row)
                if row_tuple not in unique_entries:
                    unique_entries.add(row_tuple)
                    combined_rows.append(row)
                    file_unique_count += 1
        logging.info(f"Added {file_unique_count} unique entries from {os.path.basename(csv_file)}")
    except Exception as e:
        logging.error(f"Error reading {os.path.basename(csv_file)}: {e}")

# Step 5: Convert combined rows to DataFrame
combined_df = pd.DataFrame(combined_rows, columns=[
    'Presentation / Pack Level', 'VMP / VMPP/ AMP / AMPP', 'BNF Code', 'BNF Name',
    'SNOMED Code', 'DM+D: Product Description', 'Strength', 'Unit Of Measure',
    'DM+D: Product and Pack Description', 'Pack', 'Sub-pack'
])  # Assuming 11 columns (A:K)

# Ensure all data is treated as string and fill NaN with empty strings
combined_df = combined_df.astype(str).replace('nan', '')

# Step 6: Save the combined dataframe to a single CSV file
output_csv_file = os.path.join(dirs['output'], 'combined_bnf_snomed_mapping_data.csv')
combined_df.to_csv(output_csv_file, index=False)
logging.info(f"Data combined and saved to {output_csv_file} with {len(combined_df)} rows.")

# Step 7: Convert the final CSV to XLSX using openpyxl
output_xlsx_file = os.path.join(dirs['output'], 'combined_bnf_snomed_mapping_data.xlsx')
wb = Workbook()
ws = wb.active
ws.title = "BNF Snomed Mapping Data"

# Write DataFrame to Excel sheet
for r in dataframe_to_rows(combined_df, index=False, header=True):
    ws.append(r)

# Ensure the SNOMED Code column is formatted as text
for cell in ws['F']:  # Assuming 'SNOMED Code' is in column F
    cell.number_format = '@'

# Set all column widths to 20
for col in ws.iter_cols(min_col=1, max_col=11, min_row=1, max_row=1):
    for cell in col:
        ws.column_dimensions[cell.column_letter].width = 20

# Apply the table style
tab = Table(displayName="BNFSnomedMapping", ref=ws.dimensions)
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleLight8",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True
)
ws.add_table(tab)

wb.save(output_xlsx_file)
logging.info(f"Data also saved to {output_xlsx_file}.")

# Step 8: Move the latest file to the latest directory
def extract_date_from_filename(filename):
    date_match = re.search(r'\d{8}', filename)
    if date_match:
        return datetime.strptime(date_match.group(), '%Y%m%d')
    else:
        return None

max_date = extract_date_from_filename(max(xlsx_files, key=extract_date_from_filename))
latest_files = [file for file in xlsx_files if extract_date_from_filename(file) == max_date]

if latest_files:
    latest_file = latest_files[0]
    latest_src_path = latest_file
    latest_dest_path = os.path.join(dirs['latest'], os.path.basename(latest_file))

    # Step 8.1: Check if the current latest file is outdated and remove it if necessary
    for existing_file in os.listdir(dirs['latest']):
        existing_file_path = os.path.join(dirs['latest'], existing_file)
        if os.path.isfile(existing_file_path):
            os.remove(existing_file_path)
            logging.info(f"Removed outdated file {os.path.basename(existing_file_path)}")

    # Step 8.2: Move the latest file to the latest directory
    if not os.path.exists(latest_dest_path):
        os.rename(latest_src_path, latest_dest_path)
        logging.info(f"Moved {os.path.basename(latest_src_path)} to {os.path.basename(latest_dest_path)}")
    else:
        logging.info(f"{os.path.basename(latest_dest_path)} already exists.")
